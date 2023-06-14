from PyPDF2 import PdfReader
from openpyxl import load_workbook



def common_elements(list1, list2):
    result = []
    for element in list1:
        if element in list2:
            result.append(element)
    return result

def extract_data_excel(file_path, sheet_name):
    try:
        # Load the Excel file
        workbook = load_workbook(file_path)

        # Select a specific sheet
        sheet = workbook[sheet_name]

        # Extract values from the sheet and put them into a list
        data_list = []
        for row in sheet.iter_rows(values_only=True):
            data_list.append(row)

        return data_list

    except KeyError:
        print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Usage example





def extract_text_from_pdf(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PdfReader(file)
        text = ""
        
        for page in reader.pages:
            text += page.extract_text()
        
        return text



def extract_lines_between_conditions(text, condition_start, condition_end):
    lines = text.split('\n')  # Split the text into lines
    extract = False
    extracted_lines = []

    for line in lines:
        if condition_start in line:
            extract = True  # Set the extract flag to True

        if extract:
            extracted_lines.append(line)  # Add the line to the extracted lines

        if condition_end in line:
            break  # Halt the loop if the second termination condition is met

    extracted_text = '\n'.join(extracted_lines)  # Join the extracted lines back into a single string
    return extracted_text

def extract_lines_starting_with_number(text):
    lines = text.splitlines()  # Split the text into lines
    extracted_lines = []

    for line in lines:
        if line and line[0].isdigit():  # Check if the first character is a digit
            extracted_lines.append(line)  # Add the line to the extracted lines

    extracted_text = '\n'.join(extracted_lines)  # Join the extracted lines back into a single string
    return extracted_text



def combine_lists(list1, list2):
    new_list = []
    for item1 in list1:
        name1 = item1[0]
        for item2 in list2:
            if name1 in item2:
                combined_values = item1 + tuple(item2.split(' ', 1)[1].split())
                new_list.append(combined_values)
                break
    return new_list

# Example usage:
list1 = [('2.1Maureen Patricia GUN', 'AM', '6:30-14:30', 'A', 'E', 'No'),
         ('482Pragya ADHIKARY KH', 'AM', '6:30-14:30', 'J', None, 'Yes')]

list2 = ['6:30-14:30 379Belen Ampoc MACKE', '6:30-14:30 2.1Maureen Patricia GUN']

result = combine_lists(list1, list2)
print(result)



def acacia_am(data):
    extracted_lists = []
    count = 0

    for item in data:
        if item[2] == '6:30-14:30' and item[3] == 'A':
            extracted_lists.append(item)
            count += 1

            if count == 2:
                break

    if count < 2:
        for item in data:
            if item[2] == '6:30-14:30' and item[4] == 'A':
                extracted_lists.append(item)
                count += 1

                if count == 2:
                    break

    for item in data:
        if item[2] == '7:00-13:30':
            extracted_lists.append(item)
            break

    return extracted_lists




#Passing Values to the function

pdf_path = 'timesheet_blank.pdf'
text = extract_text_from_pdf(pdf_path)


NA_all = extract_lines_between_conditions(text, "NA", "PCW")
filtered_NA_all = extract_lines_starting_with_number(NA_all)
#print("List of all the NA including AM and PM" + filtered_NA_all)


PCW_all = extract_lines_between_conditions(text, "PCW", "Physio")
filtered_PCW_all = extract_lines_starting_with_number(PCW_all)
#print("List of all the PCW including AM and PM" + filtered_PCW_all)


print('---------------------This is from pdf------------------')
all_morning_staff = filtered_NA_all + filtered_PCW_all

print('---------------------Storing string in the list------------------')
lines = all_morning_staff.splitlines()
print(lines)



file_path = 'abc.xlsx'
sheet_name = 'Sheet1'


print('---------------------This is from excel------------------')
file_path = 'abc.xlsx'
sheet_name = 'Sheet1'
data = extract_data_excel(file_path, sheet_name)
print(data)

print('---------------------common member from the list and excel------------------')

result = combine_lists( data, lines)
print(result)

print('------ acacia------')





result_acacia = acacia_am(result)
print(result_acacia)